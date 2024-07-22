from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QMessageBox, QFileDialog
import pytesseract
import PIL.Image
count = 6

class Ui_MainWindow(object):

    def Image_Scan(self):

        search = self.Line_Edit_1.text()

        if search == "":
            msg = QMessageBox()
            msg.setWindowTitle("Error !")
            msg.setText("Please Provide The file Path !")
            a = msg.exec()

        else:

            try:

                self.Label_2.setPixmap(QtGui.QPixmap(search))

                myconfig = r"--psm 11 --oem 3"
                text = pytesseract.image_to_string(
                    PIL.Image.open(search), config=myconfig)

                self.Line_Edit_2.setText(text)

            except:

                msg = QMessageBox()
                msg.setWindowTitle("Error !")
                msg.setText("Please check the file Path !")
                b = msg.exec()

    def save_Data(self):

        wb = load_workbook("Book1.xlsx")
        ws = wb.active

        File_Path = self.Line_Edit_1.text()
        text_found = self.Line_Edit_2.text()
        name_for_image = self.Line_Edit_3.text()

        if File_Path == "" or name_for_image == "":
            msg = QMessageBox()
            msg.setWindowTitle("Error !")
            msg.setText("Please Provide Required Info !")
            c = msg.exec()

        else:

            ws['B5'].value = str(name_for_image)
            ws['C5'].value = str(text_found)
            ws['D5'].value = str(File_Path)

            ws.insert_rows(5)
            wb.save("Book1.xlsx")

            self.Line_Edit_1.setText("")
            self.Line_Edit_2.setText("")
            self.Line_Edit_3.setText("")
            self.Label_2.setPixmap(QtGui.QPixmap(""))
            self.Label_2.setText("Image Will Be Shown Here")

            msg = QMessageBox()
            msg.setWindowTitle("success !")
            msg.setText("Data Uploaded To Excel Sheet Successfully !")
            d = msg.exec()

    def File_Dialog(self):

        fname = QFileDialog.getOpenFileName(None, 'Open File', 'C:\\', 'All Files (*)') 
        if fname:
            self.Line_Edit_1.setText(fname[0])
            self.Image_Scan()

    def Pevious_Img(self):

        wb = load_workbook("Book1.xlsx")
        ws = wb.active
        global count

        name_for_image = ws['B'+str(count)].value
        text_found = ws['C'+str(count)].value
        File_Path = ws['D'+str(count)].value

        if ( File_Path == None ):
            msg = QMessageBox()
            msg.setWindowTitle("Error !")
            msg.setText("No Previous Data Was Found !")
            e = msg.exec()

        else:
            self.Line_Edit_1.setText(File_Path)
            self.Line_Edit_2.setText(text_found)
            self.Line_Edit_3.setText(name_for_image)
            self.Label_2.setPixmap(QtGui.QPixmap(File_Path))
            count = count+1

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.setWindowModality(QtCore.Qt.NonModal)
        MainWindow.resize(500, 750)
        MainWindow.setMinimumSize(QtCore.QSize(500, 750))
        MainWindow.setMaximumSize(QtCore.QSize(500, 750))
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("Icon/Logo.jpg"),
                       QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainWindow.setWindowIcon(icon)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.centralwidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setSpacing(0)
        self.verticalLayout.setObjectName("verticalLayout")
        self.Main_Body = QtWidgets.QFrame(self.centralwidget)
        self.Main_Body.setStyleSheet("QFrame#Main_Body{\n"
                                     "    background-color: rgb(250, 247, 240);\n"
                                     "}\n"
                                     "QPushButton#Push_Button_1,\n"
                                     "QPushButton#Push_Button_2,\n"
                                     "QPushButton#Push_Button_3,\n"
                                     "QPushButton#Push_Button_4{\n"
                                     "    background-color: rgb(188, 206, 248);\n"
                                     "    border: 2px solid rgb(205, 252, 246);\n"
                                     "    border-radius: 17px;\n"
                                     "}\n"
                                     "QLineEdit#Line_Edit_1,\n"
                                     "QLineEdit#Line_Edit_2,\n"
                                     "QLineEdit#Line_Edit_3{\n"
                                     "    border: 3px solid rgb(205, 252, 246);\n"
                                     "    border-radius: 17px;\n"
                                     "    padding-right: 20px;\n"
                                     "    padding-left: 15px;\n"
                                     "    padding-bottom: 2px;\n"
                                     "}\n"
                                     "QLabel#Label_2{\n"
                                     "    border: 2px solid #ccc;\n"
                                     "    border-radius: 15px;\n"
                                     "}\n"
                                     "Line#line{\n"
                                     "    background-color: rgb(0, 0, 0);\n"
                                     "}")
        self.Main_Body.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.Main_Body.setFrameShadow(QtWidgets.QFrame.Raised)
        self.Main_Body.setObjectName("Main_Body")
        self.Label_1 = QtWidgets.QLabel(self.Main_Body)
        self.Label_1.setGeometry(QtCore.QRect(125, 10, 250, 50))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(16)
        font.setItalic(True)
        self.Label_1.setFont(font)
        self.Label_1.setAlignment(QtCore.Qt.AlignCenter)
        self.Label_1.setObjectName("Label_1")
        self.Label_2 = QtWidgets.QLabel(self.Main_Body)
        self.Label_2.setGeometry(QtCore.QRect(20, 200, 460, 360))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.Label_2.setFont(font)
        self.Label_2.setText("")
        self.Label_2.setPixmap(QtGui.QPixmap("pics/1-removebg-preview.png"))
        self.Label_2.setScaledContents(True)
        self.Label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.Label_2.setObjectName("Label_2")
        self.Push_Button_1 = QtWidgets.QPushButton(self.Main_Body)
        self.Push_Button_1.setGeometry(QtCore.QRect(380, 70, 100, 35))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(11)
        font.setItalic(True)
        self.Push_Button_1.setFont(font)
        self.Push_Button_1.setCursor(
            QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.Push_Button_1.setObjectName("Push_Button_1")
        self.Push_Button_1.clicked.connect(self.Image_Scan)
        self.Push_Button_4 = QtWidgets.QPushButton(self.Main_Body)
        self.Push_Button_4.setGeometry(QtCore.QRect(125, 690, 250, 35))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(12)
        font.setItalic(True)
        self.Push_Button_4.setFont(font)
        self.Push_Button_4.setCursor(
            QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.Push_Button_4.setObjectName("Push_Button_4")
        self.Push_Button_4.clicked.connect(self.save_Data)
        self.Line_Edit_1 = QtWidgets.QLineEdit(self.Main_Body)
        self.Line_Edit_1.setGeometry(QtCore.QRect(20, 70, 350, 35))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.Line_Edit_1.setFont(font)
        self.Line_Edit_1.setInputMask("")
        self.Line_Edit_1.setObjectName("Line_Edit_1")
        self.line = QtWidgets.QFrame(self.Main_Body)
        self.line.setGeometry(QtCore.QRect(0, 180, 500, 2))
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.Label_3 = QtWidgets.QLabel(self.Main_Body)
        self.Label_3.setGeometry(QtCore.QRect(20, 635, 175, 35))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(14)
        font.setItalic(True)
        self.Label_3.setFont(font)
        self.Label_3.setObjectName("Label_3")
        self.Line_Edit_3 = QtWidgets.QLineEdit(self.Main_Body)
        self.Line_Edit_3.setGeometry(QtCore.QRect(200, 635, 275, 35))
        self.Line_Edit_3.setObjectName("Line_Edit_3")
        self.Line_Edit_2 = QtWidgets.QLineEdit(self.Main_Body)
        self.Line_Edit_2.setGeometry(QtCore.QRect(25, 580, 450, 35))
        self.Line_Edit_2.setObjectName("Line_Edit_2")
        self.Push_Button_2 = QtWidgets.QPushButton(self.Main_Body)
        self.Push_Button_2.setGeometry(QtCore.QRect(90, 125, 150, 35))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(11)
        font.setItalic(True)
        self.Push_Button_2.setFont(font)
        self.Push_Button_2.setCursor(
            QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("Icon/left-arrow.png"),
                        QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.Push_Button_2.setIcon(icon1)
        self.Push_Button_2.setIconSize(QtCore.QSize(15, 15))
        self.Push_Button_2.setObjectName("Push_Button_2")
        self.Push_Button_2.clicked.connect(self.Pevious_Img)
        self.Push_Button_3 = QtWidgets.QPushButton(self.Main_Body)
        self.Push_Button_3.setGeometry(QtCore.QRect(260, 125, 150, 35))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(11)
        font.setItalic(True)
        self.Push_Button_3.setFont(font)
        self.Push_Button_3.setCursor(
            QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap("Icon/search.png"),
                        QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.Push_Button_3.setIcon(icon2)
        self.Push_Button_3.setIconSize(QtCore.QSize(15, 15))
        self.Push_Button_3.setObjectName("Push_Button_3")
        self.Push_Button_3.clicked.connect(self.File_Dialog)
        self.Label_2.raise_()
        self.Label_1.raise_()
        self.Push_Button_1.raise_()
        self.Push_Button_4.raise_()
        self.Line_Edit_1.raise_()
        self.line.raise_()
        self.Label_3.raise_()
        self.Line_Edit_3.raise_()
        self.Line_Edit_2.raise_()
        self.Push_Button_2.raise_()
        self.Push_Button_3.raise_()
        self.verticalLayout.addWidget(self.Main_Body)
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate(
            "MainWindow", "Image_Scanner_Application"))
        self.Label_1.setText(_translate("MainWindow", "Image Scanner App"))
        self.Push_Button_1.setText(_translate("MainWindow", "Search"))
        self.Push_Button_4.setText(_translate("MainWindow", "Submit"))
        self.Line_Edit_1.setPlaceholderText(_translate("MainWindow", "Search"))
        self.Label_3.setText(_translate("MainWindow", "Name Of The Image"))
        self.Line_Edit_3.setPlaceholderText(_translate(
            "MainWindow", "Give A Short Name To Identify The Image"))
        self.Line_Edit_2.setPlaceholderText(_translate(
            "MainWindow", "Text Written In The Image"))
        self.Push_Button_2.setText(_translate("MainWindow", "Previous"))
        self.Push_Button_3.setText(_translate("MainWindow", "Browse"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
